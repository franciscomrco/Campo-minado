# -*- coding: utf-8 -*-

# Beatriz Borzi
# Francisco Américo 
# Gabriel Gonçalves
# Patrícia Sparn

import os

def menu():
    opcao = 0
    
    while opcao != 9:
        limpaTerminal()
        print('---- # CAMPO MINADO # ----')
        print('------- # KABUM # --------')
        print('\n\n')
        print('ESCOLHA UMA DAS SEGUINTES OPÇÕES:')

        print('OPÇÃO 1 - ESCOLHA O TAMANHO DO TABULEIRO')
        print('OPÇÃO 2 - ESCOLHA O NIVEL DE DIFICULDADE')
        print('OPÇÃO 3 - INICIAR O JOGO')
        print('OPÇÃO 9 - SAIR')
        print('\n') 
        
        opcao = int(input('DIGITE UMA DAS OPÇÕES ACIMA: '))
        
        if opcao == 1:
            dimensao = opcao01()
        elif opcao == 2:
            dificuldade = opcao02()
        elif opcao ==3:
            try:
                conf = verificaParametros(dimensao, dificuldade)
                opcao03(conf)
            except:
                print('\nDefina as configurações do jogo antes de jogar!\n')
                input('Pressione ENTER para continuar')
        elif opcao == 9:
            print('\nGOODBYE!')
        
        print('\n') # NÃO REMOVER

def opcao01():
    while True:
        limpaTerminal()
        print('---- # CAMPO MINADO # ----')
        print('------- # KABUM # --------')
        print('\n')
        print('ESCOLHA O TAMANHO DO TABULEIRO')
        print('Quantidade minima igual a cinco')
        print('\n')

        dimensaoTabuleiro = ''
        try:
            qtLinha = int(input('Digite a quantidade de linhas: '))
            qtColuna = int(input('\nDigite a quantidade de colunas: '))
            qtLinha = 5 if qtLinha < 5 else qtLinha
            qtColuna = 5 if qtColuna < 5 else qtColuna
            print('\nVOCÊ ESCOLHEU {} LINHAS E {} COLUNAS.'.format(qtLinha, qtColuna))
            dimensaoTabuleiro = str(qtLinha) + '|' + str(qtColuna)
            input('\n\nPressione ENTER para continuar')
        except:
            print('Valor inválido')
        return dimensaoTabuleiro
        break

def opcao02():
    while True:
        limpaTerminal()
        print('---- # CAMPO MINADO # ----')
        print('------- # KABUM # --------')
        print('\n')
        print('ESCOLHA O NIVEL DE DIFICULDADE')
        print('\n')

        print('OPÇÃO 1 - FÁCIL')
        print('OPÇÃO 2 - MÉDIO')
        print('OPÇÃO 3 - DIFÍCIL')
        print('\n')

        try:
            opDificuldade = int(input('Digite o nível escolhido: '))
            if opDificuldade == 1:
                mensagem = 'BOA SORTE!'
            elif opDificuldade == 2:
                mensagem = 'CUIDADO PARA NÃO EXPLODIR!'
            elif opDificuldade == 3:
                mensagem = 'MINHA NOSSA QUE PERIGO!!'
            print('\nVOCÊ ESCOLHEU O NÍVEL {}. {}'.format(opDificuldade, mensagem))
            input('\n\nPressione ENTER para continuar')
        except:
            print('Valor inválido')
        return opDificuldade
        break

def opcao03(configuracoes):
    configuraJogo(configuracoes)
    qtLinha = configuracoes[0]
    qtColuna = configuracoes[1]
    qtMinas = configuracoes[2]
    qtCampo = (qtLinha * qtColuna) - qtMinas
    qtJogada = 1
    
    while True:
        limpaTerminal()
        print('---- # CAMPO MINADO # ----')
        print('------- # KABUM # --------')
        print('\n')
        print('      MINAS     JOGADAS')
        print('        {}         {}'.format(qtMinas, qtJogada))
        print('')
        
        # EXIBIR CAMPO TELA
        exibeCampo(qtLinha, qtColuna)
        
        linhaEscolhida = 0
        colunaEscolhida = 0
        
        if qtJogada > qtCampo:
            print('\nPARABÉNS!!')
            print('VOCÊ GANHOU!!')
            print('')
            input('Pressione ENTER para continuar')
            break

        print('') # NÃO REMOVER
        while linhaEscolhida == 0 or colunaEscolhida == 0:
            try:
                if linhaEscolhida == 0:
                    linhaEscolhida = int(input('Digite número da linha: '))
                    if linhaEscolhida < 1 or linhaEscolhida > qtLinha:
                        linhaEscolhida = 0
                if colunaEscolhida == 0:
                    colunaEscolhida = int(input('Digite número da coluna: '))
                    if colunaEscolhida < 1 or colunaEscolhida > qtColuna:
                        colunaEscolhida = 0
            except:
                input('VALOR INVÁLIDO. Pressione ENTER para continuar')
                pass
        
        # VALIDA LINHA X COLUNA ESCOLHIDA PELO USUÁRIO
        # INVÁLIDO == 0, VÁLIDO == 1
        resultado = validaLinhaColuna(linhaEscolhida, colunaEscolhida)
        
        if resultado == 0:
            print('\nJogada já efetuada, escolha outro número')
        
        # SE RESULTADO == 1 EFETUA JOGADA
        if resultado == 1:
            jogada = realizaJogada(linhaEscolhida, colunaEscolhida)
            if jogada == '*':
                print('\n       CUIDADO!!')
                print('VOCÊ ATINGIU UMA MINA!\n')
                exibeCampo(qtLinha, qtColuna)
                print('\nGAME OVER!!')
                print('')
                input('Pressione ENTER para continuar')
                break
            else:
                qtJogada += 1
        print('\n')

def verificaParametros(dimensao, dificuldade):
    parametros = dimensao.split('|')
    qtLinha = int(parametros[0])
    qtColuna = int(parametros[1])
    qtMinas = int((qtLinha * qtColuna * dificuldade) / 10)
    parametros = []
    parametros.append(qtLinha)
    parametros.append(qtColuna)
    parametros.append(qtMinas)
    return parametros

def sortearMinas(qtLinha, qtColuna, qtMinas):
    import random
    minasSorteadas =  []
    while len(minasSorteadas) < qtMinas:
        mina = (random.randint(1,qtLinha * qtColuna))
        if mina not in minasSorteadas:
            minasSorteadas.append(mina)
    minasSorteadas.sort()
    return minasSorteadas

def configuraJogo(parametros):
    from openpyxl import Workbook
    wb = Workbook()
    del wb['Sheet']
    campo_tela = wb.create_sheet('CAMPO TELA', 0)
    campo_secreto = wb.create_sheet('CAMPO SECRETO', 1)

    qtLinha = parametros[0]
    qtColuna = parametros[1]
    qtMinas = parametros[2]
    minasSorteadas = sortearMinas(qtLinha, qtColuna, qtMinas)

    wb['CAMPO TELA']
    contador = 1
    for i in range(1, qtLinha + 1):
        linha = []
        for j in range(1, qtColuna + 1):
            linha.append('-')
            contador += 1
        campo_tela.append(linha)

    wb['CAMPO SECRETO']
    contador = 1
    for i in range(1, qtLinha + 1):
        linha = []
        for j in range(1, qtColuna + 1):
            if contador in minasSorteadas:
                linha.append('*')
            else:
                linha.append('0')
            contador += 1
        campo_secreto.append(linha)
        
    campo_secreto = wb['CAMPO SECRETO']
    contador = 0
    minas = 0
    for i in range(1, qtLinha + 1):
        for j in range(1, qtColuna + 1):
            refa = campo_secreto.cell(row = i, column = j)

            try:
                if refa.value != '*':
                    ref1 = campo_secreto.cell(row = i-1, column = j-1)
                    if ref1.value == '*':
                        minas += 1
            except:
                pass
            try:
                if refa.value != '*':
                    ref2 = campo_secreto.cell(row = i-1, column = j+0)
                    if ref2.value == '*':
                        minas += 1
            except:
                pass        
            try:
                if refa.value != '*':
                    ref3 = campo_secreto.cell(row = i-1, column = j+1)
                    if ref3.value == '*':
                        minas += 1
            except:
                pass          
            try:
                if refa.value != '*':
                    ref4 = campo_secreto.cell(row = i, column = j-1)
                    if ref4.value == '*':
                        minas += 1
            except:
                pass             
            try:
                if refa.value != '*':
                    ref5 = campo_secreto.cell(row = i, column = j+1)
                    if ref5.value == '*':
                        minas += 1
            except:
                pass           
            try:
                if refa.value != '*':
                    ref6 = campo_secreto.cell(row = i+1, column = j-1)
                    if ref6.value == '*':
                        minas += 1
            except:
                pass          
            try:
                if refa.value != '*':
                    ref7 = campo_secreto.cell(row = i+1, column = j+0)
                    if ref7.value == '*':
                        minas += 1
            except:
                pass          
            try:
                if refa.value != '*':
                    ref8 = campo_secreto.cell(row = i+1, column = j+1)
                    if ref8.value == '*':
                        minas += 1
            except:
                pass             

            if minas != 0:
                if campo_secreto.cell(row = i, column = j).value != '*':
                    campo_secreto.cell(row = i, column = j).value = minas
            minas = 0
            contador += 1

    wb.save('temporario.xlsx')
    wb.close()

def exibeCampo(qtLinha, qtColuna):
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'temporario.xlsx', read_only = False)
    campo_tela = wb['CAMPO TELA']
    campo_secreto = wb['CAMPO SECRETO']
    
    numLinha = 0
    for i in range(1, qtColuna + 1):
        print('    '+ str(i), end='')
    print('') # NÃO REMOVER
    for i in range(1, qtLinha + 1):
        linha = []
        for j in range(1, qtColuna + 1):
            value = campo_tela.cell(row = i, column = j).value
            linha.append(value)
        numLinha += 1
        print(numLinha, linha)
    wb.close()   

def validaLinhaColuna(qtuLinha, qtuColuna):
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'temporario.xlsx', read_only = False)
    campo_tela = wb['CAMPO TELA']
    resultado = 0
    if campo_tela.cell(row = qtuLinha, column = qtuColuna).value == '-':
        resultado = 1
    return resultado

def realizaJogada(qtuLinha, qtuColuna):
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'temporario.xlsx', read_only = False)
    campo_tela = wb['CAMPO TELA']
    campo_secreto = wb['CAMPO SECRETO']
    valor = campo_secreto.cell(row = qtuLinha, column = qtuColuna).value
    campo_tela.cell(row = qtuLinha, column = qtuColuna).value = str(valor)
    wb.save('temporario.xlsx')
    wb.close()
    return valor

def limpaTerminal():
    os.system('cls' if os.name == 'nt' else 'clear')

# ----
menu()